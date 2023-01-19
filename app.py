# pip install pdfminer.six
from pdfminer.high_level import extract_text,extract_pages
from pdfminer.layout import LTTextContainer
import os
import locale
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# colocar o nome da empresa e cnpj para salvar txt ou excel
# criar funcao para salvar o txt ou excel



PASTA = 'rendimentos'
PLAN = 'planilhas'


def autosize(ws):
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True


def formata_numero(ws, intervalo):
    columns = ws[intervalo]

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'


def criar_plan(nome_arquivo):
    wb = Workbook()
    wb.save('planilhas' + os.sep + nome_arquivo)


def salvar_dados_plan(nome_arquivo, lista):
    wb = load_workbook('planilhas' + os.sep + nome_arquivo)
    ws = wb.active
    ws.title = 'Rendimentos'
    ws.append(['NONE', 'CPF', 'RENDIMENTOS'])
    for li in lista:
    
        ws.append(li)

    formata_numero(ws, 'C1:D10000')
    # formata_numero(ws, 'W:Z')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    

    wb.save('planilhas' + os.sep + nome_arquivo)



class Funcionario:
    def __init__(self):
        self.nome = ''
        self.CPF = ''
        self.rendimentos = 0.0
    
    def add_nome(self, nome):
        self.nome = nome.replace('\n', '')

    def add_CPF(self, CPF):
        self.CPF = CPF.replace('\n', '')

    def add_rendimentos(self, rendimentos):
        self.rendimentos = rendimentos.replace('\n', '')


def compara_textos(texto_previo, texto_arquivo, excecao=False):
    
    if texto_previo.upper() == texto_arquivo.replace('\n', '').upper():
            return True


    return False


def remover_txt():
    for txt in os.listdir(PASTA):
        nome, extensao = os.path.splitext(txt)
    
        if extensao == '.txt':
            os.remove(PASTA + os.sep + txt)

achou_nome = False

def gerar_rendimentos_funcionarios():

    for rendimento in os.listdir(PASTA):
        achou_nome = False
        # # Imprimir todo o texto contido em um PDF
        nome, extensao = os.path.splitext(rendimento)
        
        if extensao == '.pdf':
            caminho_pdf = PASTA + os.sep + rendimento
            text = extract_text(caminho_pdf)
            # print(text)
            # if 'Mexico' in text:
            #     print('Estamos no PDF historias.pdf')

            # # Salvar texto do PDF para um arquivo de texto
            arquivo_texto = PASTA + os.sep + rendimento.replace('.pdf', '.txt')
            with open(arquivo_texto,'w') as file:
                file.write(text)

            with open(arquivo_texto, 'r') as arquivo:
                lista = arquivo.readlines()
                # print(lista)
            # registros = [registro for registro in lista if registro.replace('\n', '').upper() == 'Nome completo'.upper()]
            linha = 0
            lista_funcionarios = []
            # for registro in lista:
            funcionario = False
            for i in range(len(lista)):
                if compara_textos('nome empresarial', lista[i]) and not achou_nome:
                    nome_arquivo = lista[i + 1].replace('\n','')
                    nome_arquivo = nome_arquivo + '.xlsx'
                    achou_nome = True
                if compara_textos('comprovante de rendimentos pagos e de', lista[i]):
                    # print(lista[i])
                    novo_funcionario = Funcionario()
                    funcionario = True
                    linha += 1
                if funcionario:
                    if compara_textos('nome completo', lista[i]):
                        novo_funcionario.add_nome(lista[i + 1])
                    if compara_textos('CPF', lista[i]):
                        novo_funcionario.add_CPF(lista[i + 5])
                    if compara_textos('valores em reais', lista[i]):
                        if lista[i + 1] != '\n':
                            novo_funcionario.add_rendimentos(lista[i + 1])
                
                        else:
                            novo_funcionario.add_rendimentos(lista[i + 2])

                        lista_funcionarios.append(novo_funcionario)
                        funcionario = False
            
            lista_informacoes = []
            for funcionario in lista_funcionarios:

                valor_em_reais = float(funcionario.rendimentos.replace('.', '').replace(',','.'))
                if valor_em_reais > 28559.70:
                    lista_informacoes.append([funcionario.nome, funcionario.CPF, valor_em_reais])
                    # print(f'{funcionario.nome} | {funcionario.CPF} | {funcionario.rendimentos}')


            criar_plan(nome_arquivo)
            salvar_dados_plan(nome_arquivo, lista_informacoes)

            

if __name__ == '__main__':
    try:
        os.mkdir('planilhas')
    except:
        for arquivo in os.listdir('planilhas'):
            os.remove('planilhas' + os.sep + arquivo)

    gerar_rendimentos_funcionarios()
    remover_txt()

    print('finalizado com sucesso')
    os.system('pause')